import openpyxl

import random
import json


def bulidExcel(datas, sheetName, fileName, load):
    workbook = None
    ### 判断是加载文件还是新建文件
    if load:
        workbook = openpyxl.load_workbook(fileName)
    else:
        workbook = openpyxl.Workbook()
    ### 将要操作的sheet页是否存在，不存在就新建
    ws = None
    if workbook.sheetnames.__contains__(sheetName): 
        ws = workbook[sheetName]
    else:
        ws = workbook.create_sheet(sheetName)

    for data in datas:
        ws.append(data)

    workbook.save(fileName)

def addHeadNames(datas, jsonObject):
    load = jsonObject['load']
    headNames = jsonObject['headNames']
    if load != None and load == False and headNames != None:
        datas.append(headNames)

def handlerPredefinedValue(predefinedValueDist, predefinedValueS,headNames):
    for index in range(len(predefinedValueS)):
        predefinedValue = predefinedValueS[index]
        if isinstance(predefinedValue, list):
            if len(headNames) > index and headNames[index] != None:
                predefinedValueDist[headNames[index]] = predefinedValue
        if isinstance(predefinedValue, dict):
            predefinedValueDist[predefinedValue['headName']]= predefinedValue['valueRanges']

def handlerRowData(predefinedValueDist, heandNames):
    data = {}
    for index in range(len(heandNames)):
       predefinedValue = predefinedValueDist[heandNames[index]]
       if predefinedValue != None and len(predefinedValue) > 0:
           data[index+1] = predefinedValue[random.randrange(0, len(predefinedValue))]
    return data

def buildData(buildJson):
    datas = []
    if buildJson == None and buildJson =='':
        return datas
    jsonObject = json.loads(buildJson)
    heandNames = jsonObject['headNames']

    ### 插入head
    addHeadNames(datas, jsonObject)
    
    ### 根据生成规则构造数据
    ### 处理预定义predefinedValue
    predefinedValueDist = {}
    handlerPredefinedValue(predefinedValueDist, jsonObject['predefinedValue'], heandNames)
    ### 构造数据
    for calcRule in jsonObject['calcRules']:
        custom = predefinedValueDist.copy()
        handlerPredefinedValue(custom, calcRule['predefinedValue'], heandNames)
        for i in range(calcRule['amount']):
            datas.append(handlerRowData(custom,heandNames))

    return datas


